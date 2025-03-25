import pandas as pd
import xlsxwriter as xr
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

from models.Supplier import Supplier


class ExportTool:
#Excel
    def export_supplier_to_excel(self, filename, suppliers):
        workbook = xr.Workbook(filename)
        worksheet = workbook.add_worksheet()

        worksheet.set_column('A:A', 20)
        worksheet.set_column('B:B', 20)
        worksheet.set_column('C:C', 20)
        worksheet.set_column('D:D', 20)
        worksheet.set_column('E:E', 20)


        bold = workbook.add_format({'bold': True})
        worksheet.write('A1', 'ID', bold)
        worksheet.write('B1', 'Tên', bold)
        worksheet.write('C1', 'Ngày Nhập Hàng', bold)
        worksheet.write('D1', 'Tên Sản Phẩm', bold)
        worksheet.write('E1', 'Số Lượng', bold)

        for i, sup in enumerate(suppliers, start=2):
            worksheet.write(f'A{i}', sup.id)
            worksheet.write(f'B{i}', sup.ten)
            worksheet.write(f'C{i}', sup.ngaynhaphang)
            worksheet.write(f'D{i}', sup.tensanpham)
            worksheet.write(f'E{i}', sup.soluong)

        workbook.close()


    def import_supplier_excel(self, filename):
        wb = load_workbook(filename)
        ws = wb[wb.sheetnames[0]]

        is_header = True
        suppliers = []

        for row in ws.values:
            if is_header:
                is_header = False
                continue

            id = row[0]
            ten = row[1]
            ngaynhaphang = row[2]
            tensanpham=row[3]
            soluong=row[4]
            supplier = Supplier(id,ten,ngaynhaphang,tensanpham,soluong)
            suppliers.append(supplier)

        wb.close()
        return suppliers
