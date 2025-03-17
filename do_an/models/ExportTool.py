import pandas as pd
import xlsxwriter as xr
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

from models.User import Employee


class ExportTool:
#Excel
    def export_employee_to_excel(self, filename, employees):
        workbook = xr.Workbook(filename)
        worksheet = workbook.add_worksheet()

        worksheet.set_column('A:A', 20)
        worksheet.set_column('B:B', 20)
        worksheet.set_column('C:C', 20)
        worksheet.set_column('D:D', 20)
        worksheet.set_column('E:E', 20)
        worksheet.set_column('F:F', 20)
        worksheet.set_column('G:G', 20)
        worksheet.set_column('H:H', 20)
        worksheet.set_column('I:I', 20)

        bold = workbook.add_format({'bold': True})
        worksheet.write('A1', 'Employee ID', bold)
        worksheet.write('B1', 'Employee Name', bold)
        worksheet.write('C1', 'UserName', bold)
        worksheet.write('D1', 'Password', bold)
        worksheet.write('E1', 'Role', bold)
        worksheet.write('F1', 'Email', bold)
        worksheet.write('G1', 'Level', bold)
        worksheet.write('H1', 'Shift', bold)
        worksheet.write('I1', 'Number', bold)
        worksheet.write('J1', 'Address', bold)

        for i, emp in enumerate(employees, start=2):
            worksheet.write(f'A{i}', emp.EmployeeId)
            worksheet.write(f'B{i}', emp.EmployeeName)
            worksheet.write(f'C{i}', emp.UserName)
            worksheet.write(f'D{i}', emp.Password)
            worksheet.write(f'E{i}', emp.Role)
            worksheet.write(f'F{i}', emp.Email)
            worksheet.write(f'G{i}', emp.Level)
            worksheet.write(f'H{i}', emp.Shift)
            worksheet.write(f'I{i}', emp.Number)
            worksheet.write(f'J{i}', emp.Address)

        workbook.close()


    def import_employee_excel(self, filename):
        wb = load_workbook(filename)
        ws = wb[wb.sheetnames[0]]

        is_header = True
        employees = []

        for row in ws.values:
            if is_header:
                is_header = False
                continue

            EmployeeId = row[0]
            EmployeeName = row[1]
            UserName = row[2]
            Password=row[3]
            Role=row[4]
            Email=row[5]
            Level=row[6]
            Shift=row[7]
            Number=row[8]
            Address=row[9]
            employee = Employee(EmployeeId, EmployeeName, UserName, Password,Role,Email,
                                Level,Shift,Number,Address)
            employees.append(employee)

        wb.close()
        return employees











