import pandas as pd
import xlsxwriter as xr
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

from DoAn.models.Supplier import Supplier


class ExportTool:
#Excel
    def export_employee_to_excel(self, filename, employees):
        workbook = xr.Workbook(filename)
        worksheet = workbook.add_worksheet()

        worksheet.set_column('A:A', 20)
        worksheet.set_column('B:B', 20)
        worksheet.set_column('C:C', 15)
        worksheet.set_column('D:D', 15)
        worksheet.set_column('E:E', 15)


        bold = workbook.add_format({'bold': True})
        worksheet.write('A1', 'ID', bold)
        worksheet.write('B1', 'Name', bold)
        worksheet.write('C1', 'Ngày Nhập Hàng', bold)
        worksheet.write('D1', 'Tên Sản Phẩm', bold)
        worksheet.write('E1', 'Số Lượng', bold)

        for i, emp in enumerate(employees, start=2):
            worksheet.write(f'A{i}', emp.id)
            worksheet.write(f'B{i}', emp.ten)
            worksheet.write(f'C{i}', emp.ngaynhaphang)
            worksheet.write(f'D{i}', emp.tensanpham)
            worksheet.write(f'E{i}', emp.soluong)

        workbook.close()


    def import_employee_excel(self, filename):
        wb = load_workbook(filename)
        ws = wb[wb.sheetnames[0]]

        is_header = True
        employees = []

        for row in ws.values:
            if is_header:
                is_header = False
                continue

            id = row[0]
            ten = row[1]
            ngaynhaphang = row[2]
            tensanpham=row[3]
            soluong=row[4]
            employee = Supplier(id,ten,ngaynhaphang,tensanpham,soluong)
            employees.append(employee)

        wb.close()
        return employees
