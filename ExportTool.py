import pandas as pd
import xlsxwriter as xr
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

from models.User import Employee


class ExportTool:
#Excel
    def export_employee_to_excel(self, filename, employees):
        workbook = xr.Workbook(filename)
        worksheet = workbook.add_worksheet()

        worksheet.set_column('A:A', 20)
        worksheet.set_column('B:B', 20)
        worksheet.set_column('C:C', 15)
        worksheet.set_column('D:D', 15)  # Cột Password

        bold = workbook.add_format({'bold': True})
        worksheet.write('A1', 'Employee ID', bold)
        worksheet.write('B1', 'Employee Name', bold)
        worksheet.write('C1', 'UserName', bold)
        worksheet.write('D1', 'Password', bold)

        for i, emp in enumerate(employees, start=2):
            worksheet.write(f'A{i}', emp.EmployeeId)
            worksheet.write(f'B{i}', emp.EmployeeName)
            worksheet.write(f'C{i}', emp.UserName)
            worksheet.write(f'D{i}', emp.Password)

        workbook.close()


    def import_employee_excel(self, filename):
        wb = load_workbook(filename)
        ws = wb[wb.sheetnames[0]]

        is_header = True
        employees = []

        for row in ws.values:
            if is_header:
                is_header = False
                continue

            EmployeeId = row[0]
            EmployeeName = row[1]
            UserName = row[2]
            Password=row[3]
            employee = Employee(EmployeeId, EmployeeName, UserName, Password)
            employees.append(employee)

        wb.close()
        return employees











